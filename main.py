import flet as ft
from openpyxl import Workbook, load_workbook
import os
import subprocess
import sys

def main(page: ft.Page):
    page.title = "🐾 Ficha Clínica Veterinaria 🐾"
    page.padding = 20

    # Cambiar color de tema dinámico
    colores = [ft.Colors.BLUE, ft.Colors.GREEN, ft.Colors.PINK, ft.Colors.TEAL]
    color_index = 0
    
    def cambiar_color_tema(e):
        nonlocal color_index
        color_index = (color_index + 1) % len(colores)
        page.bgcolor = colores[color_index]
        page.update()

    theme_button = ft.ElevatedButton("🎨 Cambiar Color", on_click=cambiar_color_tema)

    def go_to_main(e):
        page.go("/main")

    inicio_view = ft.View(
        route="/inicio",
        controls=[
            ft.Column(
                [
                    ft.Text("Bienvenido a Ficha Clínica Veterinaria", size=30, weight="bold"),
                    ft.ElevatedButton("Comenzar", on_click=go_to_main)
                ],
                alignment=ft.MainAxisAlignment.CENTER,
                horizontal_alignment=ft.CrossAxisAlignment.CENTER,
                expand=True
            )
        ]
    )

    campos = {
        "Nombre del dueño": ft.TextField(label="Nombre del dueño", width=400),
        "Teléfono": ft.TextField(label="Teléfono", width=400),
        "Dirección": ft.TextField(label="Dirección", width=400),
        "Nombre de la mascota": ft.TextField(label="Nombre de la mascota", width=400),
        "Color de la mascota": ft.TextField(label="Color de la mascota", width=400),
        "Especie": ft.TextField(label="Especie", width=400),
        "Raza": ft.TextField(label="Raza", width=400),
        "Edad": ft.TextField(label="Edad", width=400),
        "Peso (kg)": ft.TextField(label="Peso (kg)", width=400),
    }

    fecha = ft.DatePicker()
    page.overlay.append(fecha)
    fecha_button = ft.ElevatedButton("📅 Elegir fecha", on_click=lambda _: (setattr(fecha, "open", True), page.update()))

    anamnesis = ft.TextField(label="Anamnesis", multiline=True, min_lines=3, width=500)
    diagnostico = ft.TextField(label="Diagnóstico", multiline=True, min_lines=3, width=500)
    tratamiento = ft.TextField(label="Tratamiento", multiline=True, min_lines=3, width=500)
    resultado = ft.Text()

    btn_abrir = ft.ElevatedButton("📂 Abrir Carpeta", disabled=False)
    btn_buscar = ft.ElevatedButton("🔍 Buscar Paciente")

    def limpiar_campos():
        for campo in campos.values():
            campo.value = ""
        anamnesis.value = ""
        diagnostico.value = ""
        tratamiento.value = ""
        resultado.value = ""
        page.update()

    def guardar(e):
        datos = [campo.value for campo in campos.values()]
        nombre_mascota = campos["Nombre de la mascota"].value.strip()
        if not nombre_mascota or "" in datos or not anamnesis.value or not diagnostico.value or not tratamiento.value or not fecha.value:
            resultado.value = "⚠️ Por favor completa todos los campos."
            page.update()
            return

        try:
            documentos = os.path.expanduser("~/Documents")
            ruta_archivo = os.path.join(documentos, "Ficha_Clinica.xlsx")

            if os.path.exists(ruta_archivo):
                wb = load_workbook(ruta_archivo)
            else:
                wb = Workbook()
                # Eliminar la hoja por defecto si existe
                if "Sheet" in wb.sheetnames:
                    del wb["Sheet"]

            if nombre_mascota in wb.sheetnames:
                ws = wb[nombre_mascota]
            else:
                ws = wb.create_sheet(title=nombre_mascota)
                ws.append(list(campos.keys()) + ["Fecha", "Anamnesis", "Diagnóstico", "Tratamiento"])

            ws.append(datos + [str(fecha.value), anamnesis.value, diagnostico.value, tratamiento.value])
            wb.save(ruta_archivo)

            resultado.value = f"✅ Ficha guardada en hoja: {nombre_mascota}"
            page.update()
            limpiar_campos()

        except Exception as error:
            resultado.value = f"🚫 Error al guardar: {str(error)}"
            page.update()

    def abrir_excel(e):
        carpeta = os.path.expanduser("~/Documents")
        try:
            if sys.platform.startswith('win'):
                os.startfile(carpeta)
            elif sys.platform.startswith('darwin'):
                subprocess.call(['open', carpeta])
            else:
                subprocess.call(['xdg-open', carpeta])
        except Exception as error:
            resultado.value = f"🚫 Error al abrir carpeta: {str(error)}"
            page.update()

    def buscar_paciente(e):
        nombre = campos["Nombre de la mascota"].value.strip()
        if not nombre:
            resultado.value = "⚠️ Ingresa el nombre de la mascota."
            page.update()
            return

        try:
            documentos = os.path.expanduser("~/Documents")
            ruta_archivo = os.path.join(documentos, "Ficha_Clinica.xlsx")
            if not os.path.exists(ruta_archivo):
                resultado.value = "🚫 Archivo no encontrado."
                page.update()
                return

            wb = load_workbook(ruta_archivo)
            if nombre in wb.sheetnames:
                resultado.value = f"✅ Hoja encontrada para: {nombre}"
                abrir_excel(None)
            else:
                resultado.value = "🚫 No se encontró esa mascota."
            page.update()

        except Exception as error:
            resultado.value = f"🚫 Error al buscar: {str(error)}"
            page.update()

    btn_guardar = ft.ElevatedButton("💾 Guardar Ficha Clínica", on_click=guardar, bgcolor="green", color="white")
    btn_abrir.on_click = abrir_excel
    btn_buscar.on_click = buscar_paciente

    main_view = ft.View(
        route="/main",
        controls=[
            theme_button,
            ft.ListView(
                controls=[
                    ft.Text("🐾 Ficha Clínica Veterinaria 🐾", size=30, weight="bold", color="blue"),
                    ft.Divider(),
                    btn_buscar,
                    *campos.values(),
                    fecha_button,
                    anamnesis,
                    diagnostico,
                    tratamiento,
                    ft.Row([btn_guardar, btn_abrir], spacing=20),
                    resultado
                ],
                expand=True,
                spacing=10
            )
        ]
    )

    def route_change(e):
        if page.route == "/inicio":
            page.views.clear()
            page.views.append(inicio_view)
        elif page.route == "/main":
            page.views.clear()
            page.views.append(main_view)
        page.update()

    def view_pop(e):
        if len(page.views) > 1:
            page.views.pop()
            top_view = page.views[-1]
            page.go(top_view.route)

    page.on_route_change = route_change
    page.on_view_pop = view_pop
    page.theme_mode = ft.ThemeMode.DARK
    page.go("/inicio")

ft.app(target=main)